#!/usr/bin/env python
# _*_ coding: utf-8 _*_
import openpyxl
from datetime import datetime
import os.path
import shutil


def now_month():
    x = datetime.now()
    month = ["jan", "feb", "mar", "apr", "may", "june", "jule", "aug", "sep", "oct", "nowem",
             "dec"]
    res = str(month[x.month - 1])
    return res

windows_flag=1 # 0 если на распбери
if windows_flag == 0:
    links = ["months//"+ now_month() + ".xlsx",'templates//Образец — копия.xlsx',"months//" + now_month() + "1.xlsx"]
else:
    links= ["//home//pi//Desktop//Press_J//months//"+ now_month() + ".xlsx",'//home//pi//Desktop//Press_J//templates//Образец — копия.xlsx',"//home//pi//Desktop//Press_J//months//" + now_month() + "1.xlsx"]


def create_xl():

    students = [['Иванов Иван Владимирович','Кайка Ариана Хамаюновна','Касьян Михаил Александрович','Кпепусевич Вероника Евгеньевна','Ковалев Даниил Владимирович','Косташ Ренат Вадимович','Костяков Никита Андреевич','Линецкий Иван Алексеевич','Лузанов Никита Алексеевич','Лукманов Данил Олегович','Лялюк Борис Вадимович','Маз Никита Ильич', 'Марунчак Дмитрий Николаевич','Маслова Дарья Олеговна','Мельников Даниил Игоревич','Мешко Егор Михайлович','Михайлов Иван Романович','Ожигина Евгения Артемовна','Опарин Сергей Николаевич',"test"],['Ваня Иванов','Arina Svetotch','mishakas','Вероника','Ckibersportsmen','Tacteec','Ewokasi','Vanya Linetskiy','Nekit','aalighieri','lyalyukb','R031E','Paracelzis','Lev_Pecheniev','Дашка','Даниил Мельников','EgorMeshko','evgeniiiiiiiiia','tremplix',"ewoksicilin"], ['pahnethuem','arianakayka','mishakas','aquababe11','Ckibersportsmen','Tacteec','Ewokasi','ivaline','oldPepegas','aalighieri','lyalyukb','Zam','Paracelzis','Dimitri','mister_d_15','Messerschmitt_18','EgorMeshko','Evgeniia','Sergey Tremplix','Ewoksicilin']]


    if os.path.exists(str(links[0])):

        print("File exists")

    else:
        
        j_template = openpyxl.load_workbook(links[1])
        j_sheet=j_template.active
        for i in range(0, len(students[0])):
            j_sheet.cell(10+i,2).value=students[0][i]+" "+students[1][i]+" "+students[2][i]

        j_template.save(str(links[0]))
        print("File have created")



def mark_xl(username, obj):
    shutil.copyfile(links[0],links[2])
    journal= openpyxl.open(links[2])
    journal_sheet = journal.active
    tabled_days = [2, 8, 14, 20, 26, 32, 38, 44, 50, 56, 62, 68, 74, 80, 86, 92, 98, 104, 110, 116, 122, 128, 134, 140,
                   146, 152, 158, 164, 170, 178, 182, 188]
    today_date = datetime.now().day
    for i in range(1, 200):
        try:
            if str(username) in str(journal_sheet.cell(i,2).value):
                sstring = i;
                print(sstring,journal_sheet.cell(i,2).value)
                break
        except Exception:
            print("Ошибка имя-строка")
    colomnn = tabled_days[datetime.today().day-1]+1

    while 1:
        if journal_sheet.cell(5, colomnn).value ==None:
            journal_sheet.cell(5, colomnn).value = str(obj)
            break
        elif journal_sheet.cell(5, colomnn).value == obj:
            break
        else:
            colomnn+=1

    print("colomn", colomnn, "str", sstring)
    journal_sheet.cell(sstring,colomnn).value ="+"
    print(journal_sheet.cell(sstring,colomnn).value," ячейка")
    journal.save(links[0])
    return sstring, colomnn, obj, username
