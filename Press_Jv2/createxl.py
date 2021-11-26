#!/usr/bin/env python
# _*_ coding: utf-8 _*_
import openpyxl
from datetime import datetime
import os.path


def now_month():
    x = datetime.now()
    month = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь",
             "Декабрь"]
    res = str(month[x.month - 1])
    return res

def create_xl():

    students = [['Иванов Иван Владимирович','Кайка Ариана Хамаюновна','Касьян Михаил Александрович','Кпепусевич Вероника Евгеньевна','Ковалев Даниил Владимирович','Косташ Ренат Вадимович','Костяков Никита Андреевич','Крикун Дмитрий Вадимович','Линецкий Иван Алексеевич','Лузанов Никита Алексеевич','Лукманов Данил Олегович','Лялюк Борис Вадимович','Маз Никита Ильич','Макаров Андрей Константинович','Марунчак Дмитрий Николаевич','Маслова Дарья Олеговна','Мельников Даниил Игоревич','Мешко Егор Михайлович','Михайлов Иван Романович','Ожигина Евгения Артемовна','Опарин Сергей Николаевич',"test"],['Ваня Иванов','Arina Svetotch','Михаил Касьян','aquababe11','Ckibersportsmen','NoneNone','Ewokasi','Dmitry Dmitry','Vanya Linetskiy','Nekit','aalighieri','lyalyukb','R031E','Paracelzis','Lev_Pecheniev','Дашка','Даниил Мельников','EgorMeshko','IvanML','evgeniiiiiiiiia','tremplix',"ewoksicilin"], ['pahnethuem','Arina Svetotch','Михаил Касьян','Вероника','Даниил','none','Ewokasi','DmitriKrukun3','Vanya Linetskiy','Nekit','mika kravitz','Борис','Zam','Paracelzis','Dimitri','Дашка','Даниил Мельников','Егор Мешко','Gddizzi','Evgeniia','Sergey Tremplix','Ewoksicilin']]

    if os.path.exists(str("months//"+ now_month() + ".xlsx")):

        print("File exists")

    else:
        j_template = openpyxl.open('templates//Образец — копия.xlsx')
        j_sheet=j_template.active
        for i in range(0, len(students[0])):
            j_sheet.cell(10+i,2).value=students[0][i]+" "+students[1][i]+" "+students[2][i]

        j_template.save(str("months//" + now_month() + ".xlsx"))
        print("File have created")

def mark_xl(username, obj):
    journal= openpyxl.open("months//" + now_month() + ".xlsx")
    journal_sheet = journal.active
    tabled_days = [2, 8, 14, 20, 26, 32, 38, 44, 50, 56, 62, 68, 74, 80, 86, 92, 98, 104, 110, 116, 122, 128, 134, 140,
                   146, 152, 158, 164, 170, 178, 182, 188]
    today_date = datetime.now().day
    for i in range(1, 50):
        if str(username) in str(journal_sheet.cell(i,2).value):
            sstring = i;
            print(sstring,journal_sheet.cell(i,2).value)
            break
    colomnn = tabled_days[datetime.today().day-1]+1

    while 1:
        if journal_sheet.cell(5, colomnn).value ==None:
            journal_sheet.cell(5, colomnn).value = str(obj)
            break
        elif journal_sheet.cell(5, colomnn).value == obj:
            break
        else:
            colomnn+=1

    print("colomn", colomnn, "str", sstring)
    journal_sheet.cell(sstring,colomnn).value ="+"
    print(journal_sheet.cell(sstring,colomnn).value," ячейка")
    journal.save("months//" + now_month() + ".xlsx")
    return sstring, colomnn, obj, username

